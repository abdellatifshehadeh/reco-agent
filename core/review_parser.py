"""
Parses the 'Unmatched Side-by-Side' sheet from a previously generated report.
Detects column layout automatically so it works regardless of how many notes
columns the user has added.
"""
from openpyxl import load_workbook
import pandas as pd
import io


def parse_review_file(file_obj) -> list[dict]:
    """
    Read match groups from a user-annotated side-by-side sheet.
    file_obj can be a file path string or a BytesIO / uploaded file object.
    Returns a list of match group dicts: {ccy, match_id, fab_rows, zag_rows}
    """
    if hasattr(file_obj, 'read'):
        wb = load_workbook(io.BytesIO(file_obj.read()))
    else:
        wb = load_workbook(file_obj)

    if 'Unmatched Side-by-Side' not in wb.sheetnames:
        return []

    ws = wb['Unmatched Side-by-Side']
    col_map = _detect_columns(ws)
    if not col_map:
        return []

    def g(rd, key):
        return rd.get(col_map.get(key))

    def is_num(v):
        try: float(v); return True
        except: return False

    data_rows = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        rd = {cell.column: cell.value for cell in row}
        fab_mid = g(rd, 'fab_match')
        zag_mid = g(rd, 'zag_match')
        if fab_mid is not None and not is_num(fab_mid): fab_mid = None
        if zag_mid is not None and not is_num(zag_mid): zag_mid = None
        if fab_mid is None and zag_mid is None: continue
        desc_check = g(rd, 'fab_desc')
        if desc_check and 'Subtotal' in str(desc_check): continue

        fab_ccy = g(rd, 'fab_ccy')
        zag_ccy = g(rd, 'zag_ccy')

        data_rows.append({
            'fab_ccy':    str(fab_ccy) if fab_ccy else None,
            'zag_ccy':    str(zag_ccy) if zag_ccy else None,
            'month_fab':  g(rd, 'fab_month'),
            'fab_date':   g(rd, 'fab_date'),
            'fab_type':   g(rd, 'fab_type'),
            'fab_desc':   g(rd, 'fab_desc'),
            'fab_amt':    g(rd, 'fab_amt'),
            'fab_match':  int(float(fab_mid)) if fab_mid is not None else None,
            'zag_month':  g(rd, 'zag_month'),
            'zag_date':   g(rd, 'zag_date'),
            'zag_ref':    g(rd, 'zag_ref'),
            'zag_desc':   g(rd, 'zag_desc'),
            'zag_amt':    g(rd, 'zag_amt'),
            'zag_jnl':    g(rd, 'zag_jnl'),
            'zag_match':  int(float(zag_mid)) if zag_mid is not None else None,
        })

    df = pd.DataFrame(data_rows)
    fab_rows_all = df[df['fab_match'].notna() & df['fab_ccy'].notna()].copy()
    zag_rows_all = df[df['zag_match'].notna() & df['zag_ccy'].notna()].copy()

    fab_keys = set(zip(fab_rows_all['fab_ccy'], fab_rows_all['fab_match']))
    zag_keys = set(zip(zag_rows_all['zag_ccy'], zag_rows_all['zag_match']))
    all_keys = sorted(fab_keys | zag_keys)

    groups = []
    for ccy, mid in all_keys:
        fab_rows = fab_rows_all[
            (fab_rows_all['fab_ccy'] == ccy) & (fab_rows_all['fab_match'] == mid)
        ].to_dict('records')
        zag_rows = zag_rows_all[
            (zag_rows_all['zag_ccy'] == ccy) & (zag_rows_all['zag_match'] == mid)
        ].to_dict('records')
        groups.append({
            'ccy': ccy,
            'match_id': mid,
            'fab_rows': fab_rows,
            'zag_rows': zag_rows,
        })

    return groups


def _detect_columns(ws) -> dict:
    """Detect column positions from the header row."""
    col_map = {}
    for hr in [3, 4, 5]:
        for cell in ws[hr]:
            v = str(cell.value or '').strip().lower()
            c = cell.column
            if v in ('month',) and c <= 8 and 'fab_month' not in col_map:
                col_map['fab_month'] = c
            elif v in ('ccy',) and c <= 8 and 'fab_ccy' not in col_map:
                col_map['fab_ccy'] = c
            elif any(x in v for x in ('cust date', 'fab date', 'cust settle')):
                col_map['fab_date'] = c
            elif any(x in v for x in ('cust type', 'fab type')):
                col_map['fab_type'] = c
            elif any(x in v for x in ('cust description', 'fab description')):
                col_map['fab_desc'] = c
            elif any(x in v for x in ('cust amount', 'fab amount')):
                col_map['fab_amt'] = c
            elif 'match id' in v and c <= 9:
                col_map['fab_match'] = c
            elif v in ('month',) and c >= 9 and 'zag_month' not in col_map:
                col_map['zag_month'] = c
            elif v in ('ccy',) and c >= 9 and 'zag_ccy' not in col_map:
                col_map['zag_ccy'] = c
            elif 'zag date' in v:
                col_map['zag_date'] = c
            elif 'zag ref' in v:
                col_map['zag_ref'] = c
            elif 'zag description' in v:
                col_map['zag_desc'] = c
            elif 'zag amount' in v:
                col_map['zag_amt'] = c
            elif 'zag journal' in v:
                col_map['zag_jnl'] = c
            elif 'match id' in v and c >= 9:
                col_map['zag_match'] = c
        if len(col_map) >= 8:
            break
    return col_map
