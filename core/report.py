"""
Excel report writer.
Produces the multi-tab reconciliation workbook.
"""
import io
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
GREEN   = PatternFill("solid", fgColor="C6EFCE")
AMBER   = PatternFill("solid", fgColor="FFEB9C")
RED     = PatternFill("solid", fgColor="FFC7CE")
GREY    = PatternFill("solid", fgColor="F2F2F2")
BLUE    = PatternFill("solid", fgColor="DEEAF1")
HDR     = PatternFill("solid", fgColor="1F4E79")
MHDR    = PatternFill("solid", fgColor="2E4057")
MANUAL  = PatternFill("solid", fgColor="E2EFDA")
LORANGE = PatternFill("solid", fgColor="FCE4D6")
RED_C   = PatternFill("solid", fgColor="FCE4D6")
RED_Z   = PatternFill("solid", fgColor="FFE4E1")
HDR_C   = PatternFill("solid", fgColor="C00000")
HDR_Z   = PatternFill("solid", fgColor="833C00")
HDR_MID = PatternFill("solid", fgColor="404040")
G_ALT   = PatternFill("solid", fgColor="F2F2F2")
WHITE   = PatternFill("solid", fgColor="FFFFFF")
LGREEN  = PatternFill("solid", fgColor="EBF5E0")
LBLUE   = PatternFill("solid", fgColor="E0EDF5")
HF  = Font(bold=True, color="FFFFFF", size=10)
N10 = Font(size=10)
N9  = Font(size=9)


def _hdr(ws, cols, row=1):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row, c, h)
        cell.fill = HDR; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def _fd(v):
    if v is None: return ''
    try:
        ts = pd.Timestamp(v)
        return '' if pd.isnull(ts) else ts.strftime('%d/%m/%Y')
    except Exception:
        return ''


def _fn(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return ''
    try:
        return round(float(v), 2)
    except Exception:
        return ''


def build_report(
    recon_df:      pd.DataFrame,
    balance_df:    pd.DataFrame,
    custodian_df:  pd.DataFrame,
    zag_raw_df:    pd.DataFrame,
    zag_netted_df: pd.DataFrame,
    manual_groups: list[dict],
    period_label:  str,
    custodian_name: str,
    manual_df: pd.DataFrame | None = None,
) -> bytes:
    """Build the full Excel workbook and return as bytes (for Streamlit download)."""
    wb = Workbook()

    engine_matched = recon_df[recon_df['match_status'] == 'MATCHED'].copy()
    engine_breaks  = recon_df[recon_df['match_status'].str.startswith('UNMATCHED')].copy()
    flagged_ids    = _find_flagged(manual_groups)

    # ── TAB 1: Summary ────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = 'Period Summary'
    ws1['A1'] = f'Cash Reconciliation — {custodian_name} vs ZagTrader  |  {period_label}'
    ws1['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws1['A2'] = f'Generated: {datetime.today().strftime("%d/%m/%Y %H:%M")}  |  Engine: gross+tax netting · coupon splits · income delay 15d'
    ws1['A2'].font = Font(size=10, color='595959')
    ws1.append([])

    mH = len(engine_matched[engine_matched['confidence'] == 'HIGH'])
    mM = len(engine_matched[engine_matched['confidence'] != 'HIGH'])
    mMn = len(manual_df) if manual_df is not None else 0
    uC  = len(engine_breaks[engine_breaks['match_status'] == 'UNMATCHED - CUSTODIAN ONLY'])
    uZ  = len(engine_breaks[engine_breaks['match_status'] == 'UNMATCHED - ZAG ONLY'])
    tot = mH + mM + mMn + uC + uZ

    ws1.append(['MATCHING SUMMARY']); ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=11)
    ws1.append([])
    _hdr(ws1, ['Metric', 'Count', '% of Total'], row=ws1.max_row + 1)
    for label, cnt, fill in [
        ('Total items reviewed', tot, GREY),
        ('Engine matched — high confidence', mH, GREEN),
        ('Engine matched — medium confidence', mM, AMBER),
        ('Manually confirmed by user', mMn, MANUAL),
        ('Unmatched — custodian only', uC, RED if uC else GREEN),
        ('Unmatched — ZagTrader only', uZ, RED if uZ else GREEN),
    ]:
        rn = ws1.max_row + 1
        pct = f'{round(cnt/tot*100, 1) if tot else 0}%'
        for c, v in enumerate([label, cnt, pct], 1):
            cell = ws1.cell(rn, c, v); cell.fill = fill; cell.font = N10

    ws1.append([]); ws1.append([])

    # Flagged manual groups
    if flagged_ids:
        ws1.append(['⚠ FLAGGED MANUAL MATCHES — LARGE DIFFERENCE — VERIFY'])
        ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=11, color='C00000')
        ws1.append([])
        _hdr(ws1, ['Match ID', 'CCY', 'Cust Amount', 'ZAG Amount', 'Diff', '% Off'], row=ws1.max_row + 1)
        if manual_df is not None:
            for _, r in manual_df[manual_df['confidence'] == 'REVIEW REQUIRED'].iterrows():
                rn = ws1.max_row + 1
                net = abs(r['difference']) if r.get('difference') else 0
                mx  = max(abs(r.get('cust_amount') or 1), abs(r.get('zag_amount') or 1))
                for c, v in enumerate([r['match_method'], r['currency'],
                                        _fn(r.get('cust_amount')), _fn(r.get('zag_amount')),
                                        _fn(r.get('difference')), f'{round(net/mx*100,1)}%'], 1):
                    cell = ws1.cell(rn, c, v); cell.fill = LORANGE; cell.font = Font(bold=True, size=10)
        ws1.append([]); ws1.append([])

    ws1.append(['MONTH-END CLOSING BALANCE COMPARISON']); ws1.cell(ws1.max_row, 1).font = Font(bold=True, size=11)
    ws1.append([])
    _hdr(ws1, ['Currency', 'Custodian Opening', 'Custodian Closing',
               'ZAG Opening', 'ZAG Closing', 'Difference', 'Status'], row=ws1.max_row + 1)
    for _, r in balance_df[~balance_df['status'].eq('ONE SIDE MISSING')].sort_values('currency').iterrows():
        rn = ws1.max_row + 1
        fill = GREEN if r['status'] == 'OK' else RED
        for c, v in enumerate([r['currency'], _fn(r.get('cust_opening')), _fn(r.get('cust_closing')),
                                _fn(r.get('zag_opening')), _fn(r.get('zag_closing')),
                                _fn(r.get('difference')), r['status']], 1):
            cell = ws1.cell(rn, c, v); cell.fill = fill; cell.font = N10
            cell.alignment = Alignment(horizontal='right' if c in [2, 3, 4, 5, 6] else 'center')
    for c, w in enumerate([12, 20, 20, 18, 18, 14, 12], 1):
        ws1.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 2: Matched ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Matched')
    _hdr(ws2, ['CCY', 'Confidence', 'Method', 'Cust Date', 'Cust Type', 'Cust Description',
               'Cust Amount', 'Cust Balance', 'ZAG Date', 'ZAG Ref', 'ZAG Description',
               'ZAG Amount', 'ZAG Balance', 'ZAG Journal', 'Netted?', 'Nett Type',
               'ZAG Gross', 'ZAG WHT', 'Diff', 'Notes'])
    for _, r in engine_matched.iterrows():
        fill = GREEN if r['confidence'] == 'HIGH' else AMBER
        vals = [r['currency'], r['confidence'], r['match_method'],
                _fd(r.get('cust_settle_date')), r.get('cust_type'), r.get('cust_description'),
                _fn(r.get('cust_amount')), _fn(r.get('cust_balance')),
                _fd(r.get('zag_date')), r.get('zag_ref'), r.get('zag_description'),
                _fn(r.get('zag_amount')), _fn(r.get('zag_balance')), r.get('zag_journal'),
                'YES' if r.get('zag_netted') else '', r.get('nett_type', ''),
                _fn(r.get('zag_gross')), _fn(r.get('zag_tax')),
                _fn(r.get('difference')), r.get('notes', '')]
        rn = ws2.max_row + 1
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(rn, c, v); cell.fill = fill; cell.font = N9
            cell.alignment = Alignment(horizontal='right' if c in [7, 8, 12, 13, 17, 18, 19] else 'left')
    if manual_df is not None:
        for _, r in manual_df.iterrows():
            fill = LORANGE if r.get('confidence') == 'REVIEW REQUIRED' else MANUAL
            vals = [r.get('currency'), r.get('confidence'), r.get('match_method'),
                    _fd(r.get('cust_settle_date')), r.get('cust_type'), r.get('cust_description'),
                    _fn(r.get('cust_amount')), '',
                    _fd(r.get('zag_date')), r.get('zag_ref'), r.get('zag_description'),
                    _fn(r.get('zag_amount')), '', r.get('zag_journal'),
                    '', '', '', '', _fn(r.get('difference')), r.get('notes', '')]
            rn = ws2.max_row + 1
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(rn, c, v); cell.fill = fill; cell.font = N9
    for c, w in enumerate([6, 10, 30, 14, 20, 40, 13, 14, 14, 12, 40, 13, 14, 16, 7, 14, 13, 12, 10, 50], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 3: Breaks ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Breaks')
    ws3['A1'] = 'UNMATCHED ITEMS'; ws3['A1'].font = Font(bold=True, size=12, color='C00000')
    ws3['A2'] = 'Amber = confirmed break with user note. Red = unresolved.'
    ws3['A2'].font = Font(size=9, italic=True, color='595959')
    ws3.append([])
    _hdr(ws3, ['CCY', 'Status', 'Cust Date', 'Cust Type', 'Cust Description', 'Cust Amount',
               'ZAG Date', 'ZAG Ref', 'ZAG Description', 'ZAG Amount', 'ZAG Journal',
               'Netted?', 'Notes'], row=4)
    for _, r in engine_breaks.sort_values(['currency', 'match_status']).iterrows():
        note = str(r.get('notes', ''))
        has_note = any(n in note for n in ['Confirmed', 'Outside', 'duplicate', 'scope', 'posting'])
        fill = AMBER if has_note else RED
        vals = [r['currency'], r['match_status'],
                _fd(r.get('cust_settle_date')), r.get('cust_type'), r.get('cust_description'),
                _fn(r.get('cust_amount')),
                _fd(r.get('zag_date')), r.get('zag_ref'), r.get('zag_description'),
                _fn(r.get('zag_amount')), r.get('zag_journal'),
                'YES' if r.get('zag_netted') else '', note]
        rn = ws3.max_row + 1
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(rn, c, v); cell.fill = fill; cell.font = N9
            cell.alignment = Alignment(horizontal='right' if c in [6, 10] else 'left')
    for c, w in enumerate([6, 24, 14, 20, 40, 13, 14, 12, 40, 13, 16, 7, 50], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 4: Unmatched Side-by-Side ─────────────────────────────────────────
    ws4 = wb.create_sheet('Unmatched Side-by-Side')
    ws4['A1'] = 'UNMATCHED — SIDE-BY-SIDE (add match IDs in col G and col P, then re-upload)'
    ws4['A1'].font = Font(bold=True, size=12, color='1F4E79')
    ws4.append([]); ws4.append([])
    HDR_ROW = 3
    for c, h in enumerate(['Month', 'CCY', 'Cust Date', 'Cust Type', 'Cust Description', 'Cust Amount', 'Match ID', 'Notes'], 1):
        cell = ws4.cell(HDR_ROW, c, h); cell.fill = HDR_C; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws4.cell(HDR_ROW, 9, '').fill = HDR_MID
    for c, h in enumerate(['Month', 'CCY', 'ZAG Date', 'ZAG Ref', 'ZAG Description', 'ZAG Amount', 'ZAG Journal', 'Match ID', 'Notes'], 10):
        cell = ws4.cell(HDR_ROW, c, h); cell.fill = HDR_Z; cell.font = HF
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    cust_only = engine_breaks[engine_breaks['match_status'] == 'UNMATCHED - CUSTODIAN ONLY'].copy()
    zag_only  = engine_breaks[engine_breaks['match_status'] == 'UNMATCHED - ZAG ONLY'].copy()
    cust_only['_d'] = pd.to_datetime(cust_only['cust_settle_date'], errors='coerce')
    zag_only['_d']  = pd.to_datetime(zag_only['zag_date'], errors='coerce')
    cust_only = cust_only.sort_values(['currency', '_d']).reset_index(drop=True)
    zag_only  = zag_only.sort_values(['currency', '_d']).reset_index(drop=True)

    for ccy in sorted(set(cust_only['currency'].unique()) | set(zag_only['currency'].unique())):
        fc = cust_only[cust_only['currency'] == ccy].reset_index(drop=True)
        zc = zag_only[zag_only['currency'] == ccy].reset_index(drop=True)
        if len(fc) == 0 and len(zc) == 0: continue
        rn = ws4.max_row + 1
        cell = ws4.cell(rn, 1, f'Currency: {ccy}  —  Custodian only: {len(fc)}  |  ZAG only: {len(zc)}')
        cell.fill = MHDR; cell.font = Font(bold=True, color='FFFFFF', size=10)
        ws4.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=18)
        for i in range(max(len(fc), len(zc))):
            rn = ws4.max_row + 1; fa = G_ALT if i % 2 == 0 else WHITE
            if i < len(fc):
                r = fc.iloc[i]
                vf = ['', r['currency'], _fd(r.get('cust_settle_date')),
                      str(r.get('cust_type') or '')[:25], str(r.get('cust_description') or '')[:55],
                      _fn(r.get('cust_amount')), '', '']
                ff = RED_C
            else:
                vf = ['', '', '', '', '', '', '', '']; ff = fa
            for c, v in enumerate(vf, 1):
                cell = ws4.cell(rn, c, v); cell.fill = ff; cell.font = N9
                cell.alignment = Alignment(horizontal='right' if c == 6 else 'left')
            ws4.cell(rn, 9, '').fill = HDR_MID
            if i < len(zc):
                r = zc.iloc[i]
                vz = ['', r['currency'], _fd(r.get('zag_date')), str(r.get('zag_ref') or ''),
                      str(r.get('zag_description') or '')[:55], _fn(r.get('zag_amount')),
                      str(r.get('zag_journal') or ''), '', '']
                zf = RED_Z
            else:
                vz = ['', '', '', '', '', '', '', '', '']; zf = fa
            for c, v in enumerate(vz, 10):
                cell = ws4.cell(rn, c, v); cell.fill = zf; cell.font = N9
                cell.alignment = Alignment(horizontal='right' if c == 15 else 'left')
        ft = fc['cust_amount'].apply(lambda x: float(x) if x and not (isinstance(x, float) and np.isnan(x)) else 0).sum()
        zt = zc['zag_amount'].apply(lambda x: float(x) if x and not (isinstance(x, float) and np.isnan(x)) else 0).sum()
        rn = ws4.max_row + 1
        ws4.cell(rn, 5, f'Subtotal {ccy} Custodian').font = Font(bold=True, size=9)
        ws4.cell(rn, 6, round(ft, 2)).font = Font(bold=True, size=9)
        ws4.cell(rn, 9, '').fill = HDR_MID
        ws4.cell(rn, 14, f'Subtotal {ccy} ZAG').font = Font(bold=True, size=9)
        ws4.cell(rn, 15, round(zt, 2)).font = Font(bold=True, size=9)
        ws4.append([])
    for c, w in enumerate([10, 6, 12, 20, 45, 14, 10, 22, 2, 10, 6, 12, 14, 45, 14, 16, 10, 22], 1):
        ws4.column_dimensions[get_column_letter(c)].width = w
    ws4.freeze_panes = 'A4'

    # ── TAB 5: Manual Match Detail ────────────────────────────────────────────
    if manual_groups:
        ws5 = wb.create_sheet('Manual Match Detail')
        ws5['A1'] = 'MANUAL MATCH GROUPS — User confirmed'; ws5['A1'].font = Font(bold=True, size=11, color='1F4E79')
        ws5['A2'] = 'Green = custodian lines. Blue = ZAG lines. Orange = flagged (large diff).'; ws5['A2'].font = Font(size=9, italic=True, color='595959')
        ws5.append([])
        for g in sorted(manual_groups, key=lambda x: (x['ccy'], str(x['match_id']))):
            is_f = g['match_id'] in flagged_ids
            fab_t = sum(float(r['fab_amt']) for r in g.get('fab_rows', []) if r.get('fab_amt') is not None)
            zag_t = sum(float(r['zag_amt']) for r in g.get('zag_rows', []) if r.get('zag_amt') is not None)
            rn = ws5.max_row + 1
            cell = ws5.cell(rn, 1, f"[{g['ccy']}] ID {g['match_id']}  —  {len(g.get('fab_rows',[]))} cust  |  {len(g.get('zag_rows',[]))} ZAG{'  ⚠ LARGE DIFF' if is_f else ''}")
            cell.fill = LORANGE if is_f else MHDR; cell.font = Font(bold=True, color='FFFFFF', size=10)
            ws5.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=9)
            for r in g.get('fab_rows', []):
                rn = ws5.max_row + 1
                for c, v in enumerate(['CUST', g['ccy'], '', _fd(r.get('fab_date')), str(r.get('fab_type',''))[:20], str(r.get('fab_desc',''))[:55], _fn(r.get('fab_amt')), '', ''], 1):
                    cell = ws5.cell(rn, c, v); cell.fill = LORANGE if is_f else LGREEN; cell.font = N9
            for r in g.get('zag_rows', []):
                rn = ws5.max_row + 1
                for c, v in enumerate(['ZAG', g['ccy'], '', _fd(r.get('zag_date')), str(r.get('zag_ref',''))[:20], str(r.get('zag_desc',''))[:55], _fn(r.get('zag_amt')), str(r.get('zag_jnl',''))[:15], ''], 1):
                    cell = ws5.cell(rn, c, v); cell.fill = LORANGE if is_f else LBLUE; cell.font = N9
            rn = ws5.max_row + 1
            ws5.cell(rn, 6, 'NET DIFF:').font = Font(bold=True, size=9)
            ws5.cell(rn, 7, round(fab_t + zag_t, 2)).font = Font(bold=True, size=9, color='C00000' if is_f else '000000')
            ws5.append([])
        for c, w in enumerate([6, 6, 10, 12, 22, 55, 13, 16, 10], 1):
            ws5.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 6: ZAG Netting Detail ─────────────────────────────────────────────
    ws6 = wb.create_sheet('ZAG Netting Detail')
    ws6['A1'] = 'ZagTrader: Multi-line journal entries collapsed before matching'; ws6['A1'].font = Font(bold=True, size=11, color='1F4E79')
    _hdr(ws6, ['Currency', 'Date', 'Journal Ref', 'Nett Type', 'Description',
               'Gross Amt', 'WHT Tax', 'Net Amt', 'Component Lines'], row=3)
    for _, r in zag_netted_df[zag_netted_df['netted'] == True].iterrows():
        d = r.get('debit') or 0; cr = r.get('credit') or 0
        rn = ws6.max_row + 1
        for c, v in enumerate([r['currency'], _fd(r['trade_date']), r['journal_ref'],
                                r.get('nett_type', 'gross+tax'), r['description'],
                                _fn(r.get('zag_gross')), _fn(r.get('zag_tax')),
                                _fn(d if d else cr),
                                str(r.get('component_descs', ''))[:120]], 1):
            cell = ws6.cell(rn, c, v); cell.fill = BLUE; cell.font = N9
    for c, w in enumerate([8, 12, 16, 18, 40, 14, 12, 16, 120], 1):
        ws6.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 7: Raw Custodian ──────────────────────────────────────────────────
    ws7 = wb.create_sheet('Raw - Custodian')
    _hdr(ws7, ['Custodian', 'Currency', 'Trade Date', 'Settle Date', 'Txn Type', 'Description', 'Debit', 'Credit', 'Balance'])
    for _, r in custodian_df.iterrows():
        ws7.append([r.get('custodian', ''), r['currency'], _fd(r.get('trade_date')), _fd(r.get('settlement_date')),
                    r.get('txn_type'), r.get('description'), _fn(r.get('debit')), _fn(r.get('credit')), _fn(r.get('balance'))])
    for c, w in enumerate([25, 8, 12, 12, 22, 40, 13, 13, 14], 1):
        ws7.column_dimensions[get_column_letter(c)].width = w

    # ── TAB 8: Raw ZAG ────────────────────────────────────────────────────────
    ws8 = wb.create_sheet('Raw - ZagTrader')
    _hdr(ws8, ['Currency', 'Date', 'Ref', 'Description', 'Debit', 'Credit', 'Balance', 'Journal Ref', 'Tax Line?'])
    for _, r in zag_raw_df.iterrows():
        ws8.append([r['currency'], _fd(r.get('trade_date')), r.get('ref'), r['description'],
                    _fn(r.get('debit')), _fn(r.get('credit')), _fn(r.get('balance')),
                    r.get('journal_ref'), 'YES' if r.get('is_tax_line') else ''])
    for c, w in enumerate([8, 12, 12, 60, 13, 13, 14, 16, 8], 1):
        ws8.column_dimensions[get_column_letter(c)].width = w

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _find_flagged(match_groups: list[dict]) -> set:
    flagged = set()
    for g in match_groups:
        fab_t = sum(float(r['fab_amt']) for r in g.get('fab_rows', []) if r.get('fab_amt') is not None)
        zag_t = sum(float(r['zag_amt']) for r in g.get('zag_rows', []) if r.get('zag_amt') is not None)
        if g.get('fab_rows') and g.get('zag_rows'):
            net = abs(fab_t + zag_t)
            mx  = max(abs(fab_t), abs(zag_t), 0.01)
            if net > 100 and net / mx > 0.01:
                flagged.add(g['match_id'])
    return flagged
