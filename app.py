"""
Cash Reconciliation Agent — single file version
Upload to GitHub as-is with requirements.txt, no subfolders needed.
"""
import io, re, os, sys
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Cash Reconciliation Agent", page_icon="🔄", layout="wide")

# ══════════════════════════════════════════════════════════════════════════════
# 1. FAB PARSER
# ══════════════════════════════════════════════════════════════════════════════
DEBIT_TYPES  = ['Receive vs Payment','Cash Withdrawal','Cash Out']
CREDIT_TYPES = ['Deliver vs Payment','Cash Deposit','CA Dividend','CA Income','CA Pay due']

def parse_fab(filepath):
    xls = pd.ExcelFile(filepath, engine="xlrd")
    df_raw = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=None)
    rows=[]; current_currency=None; in_txn_block=False; balance_info={}
    for _, row in df_raw.iterrows():
        vals=[str(v).strip() for v in row if str(v).strip() not in ['nan','NaN','']]
        if not vals: continue
        for v in vals:
            m=re.match(r'^Currency:\s*([A-Z]{3})$',v)
            if m: current_currency=m.group(1); in_txn_block=False; break
        if len(vals)>=3 and vals[0]=='Trade Date' and vals[1]=='Settlement Date':
            in_txn_block=True; continue
        if not in_txn_block or not current_currency: continue
        if 'Starting Balance' in vals:
            try: balance_info.setdefault(current_currency,{})['starting']=float(str(vals[-1]).replace(',',''))
            except: pass
            continue
        if 'Ending Balance' in vals or 'Current Balance' in vals:
            try: balance_info.setdefault(current_currency,{})['ending']=float(str(vals[-1]).replace(',',''))
            except: pass
            continue
        if len(vals)>=5:
            try:
                trade_date=pd.to_datetime(vals[0],dayfirst=True,errors='coerce')
                if pd.isnull(trade_date): continue
                settlement_date=pd.to_datetime(vals[1],dayfirst=True,errors='coerce')
                txn_type=vals[2]; description=vals[3]
                amount=float(str(vals[4]).replace(',','')); balance=float(str(vals[-1]).replace(',',''))
                if any(d in txn_type for d in DEBIT_TYPES): debit,credit=amount,None
                elif any(c in txn_type for c in CREDIT_TYPES): debit,credit=None,amount
                else: debit,credit=amount,None
                rows.append({'source':'CUSTODIAN','custodian':'FAB','currency':current_currency,
                             'trade_date':trade_date,'settlement_date':settlement_date,
                             'txn_type':txn_type,'description':description,
                             'debit':debit,'credit':credit,'balance':balance})
            except: continue
    return pd.DataFrame(rows), balance_info

# ══════════════════════════════════════════════════════════════════════════════
# 2. ZAGTRADER PARSER + NETTING
# ══════════════════════════════════════════════════════════════════════════════
def parse_zag_raw(filepath):
    xl=pd.ExcelFile(filepath)
    currency_sheets=[s for s in xl.sheet_names if s!='Cash Balance']
    rows=[]; balance_info={}
    for sheet in currency_sheets:
        df=pd.read_excel(xl,sheet_name=sheet,header=None)
        for i,row in df.iterrows():
            if i<4: continue
            raw=list(row)
            if str(raw[3]).strip()=='Open Balance':
                try: balance_info.setdefault(sheet,{})['starting']=float(raw[7])
                except: pass
                continue
            date_val=pd.to_datetime(str(raw[0]).strip(),format='%d %b %Y',errors='coerce')
            if pd.isnull(date_val): continue
            def to_num(v):
                try: f=float(v); return None if np.isnan(f) else f
                except: return None
            ref=str(raw[2]).strip() if str(raw[2]).strip() not in ['nan',''] else None
            desc=str(raw[3]).strip() if str(raw[3]).strip() not in ['nan',''] else ''
            debit=to_num(raw[5]); credit=to_num(raw[6])
            balance=to_num(raw[7])
            jref=str(raw[8]).strip() if str(raw[8]).strip() not in ['nan',''] else None
            if debit is not None or credit is not None:
                rows.append({'source':'ZAG','currency':sheet,'trade_date':date_val,
                             'settlement_date':date_val,'ref':ref,'description':desc,
                             'debit':debit,'credit':credit,'balance':balance,
                             'journal_ref':jref,'is_tax_line':False,'netted':False})
    try:
        df_cb=pd.read_excel(xl,sheet_name='Cash Balance',header=None,skiprows=1)
        for _,row in df_cb.iterrows():
            try:
                ccy=str(row[0]).strip(); bal=float(row[1])
                if len(ccy)==3: balance_info.setdefault(ccy,{})['ending']=bal
            except: pass
    except: pass
    return pd.DataFrame(rows), balance_info

def net_zag(zag_raw_df):
    df=zag_raw_df.copy()
    df['is_tax_line']=(df['description'].str.startswith('TAX for:',na=False)|
                       df['description'].str.startswith('Tax for:',na=False))
    netted_rows=[]; used=set()
    grouped=df[df['journal_ref'].notna()].groupby(['currency','journal_ref'])
    for (ccy,jref),grp in grouped:
        idx_list=list(grp.index)
        if len(grp)<2: continue
        gross_rows=grp[~grp['is_tax_line']]; tax_rows=grp[grp['is_tax_line']]
        if len(tax_rows)>0:
            gd=gross_rows['debit'].fillna(0).sum(); gc=gross_rows['credit'].fillna(0).sum()
            tc=tax_rows['credit'].fillna(0).sum(); td=tax_rows['debit'].fillna(0).sum()
            nd=gd-tc; nc=gc-td; base=gross_rows.iloc[0]
            netted_rows.append({**_base(base,ccy,jref),
                'debit':round(nd,4) if nd>0.001 else None,
                'credit':round(nc,4) if nc>0.001 else None,
                'balance':_lastbal(grp),'zag_gross':gd if gd else gc,
                'zag_tax':tc if tc else td,'nett_type':'gross+tax',
                'component_count':len(grp),'component_descs':' | '.join(grp['description'].tolist())})
            for i in idx_list: used.add(i)
            continue
        descs=gross_rows['description'].tolist(); fd=str(descs[0]).upper()[:50]
        all_same=all(str(d).upper()[:50]==fd for d in descs)
        def ra(r):
            d=r['debit']; c=r['credit']
            if d is not None and not(isinstance(d,float) and np.isnan(d)): return abs(float(d))
            if c is not None and not(isinstance(c,float) and np.isnan(c)): return abs(float(c))
            return 0.0
        sa=sorted([ra(r) for _,r in gross_rows.iterrows()],reverse=True)
        is_charge=(len(sa)==2 and sa[0]>0 and sa[1]<=100 and sa[0]/max(sa[1],0.01)>100)
        du=[str(d).upper() for d in gross_rows['description']]
        is_fd=(any('SALE OF SECURITY' in d or 'BOND MATURED' in d for d in du) and
               any('INTEREST' in d for d in du))
        if all_same or is_charge or is_fd:
            ad=gross_rows['debit'].fillna(0).sum(); ac=gross_rows['credit'].fillna(0).sum()
            nd=round(ad,4) if ad>0.001 else None; nc=round(ac,4) if ac>0.001 else None
            if is_charge:
                main=sa[0]
                mr=max(gross_rows.iterrows(),key=lambda x:ra(x[1]))[1]
                if mr['debit'] is not None and not(isinstance(mr['debit'],float) and np.isnan(mr['debit'])):
                    nd=round(main,4); nc=None
                else:
                    nc=round(main,4); nd=None
            nt='coupon_split' if all_same else 'transfer_charge' if is_charge else 'fd_split'
            base=gross_rows.iloc[0]
            netted_rows.append({**_base(base,ccy,jref),
                'debit':nd,'credit':nc,'balance':_lastbal(grp),
                'zag_gross':ad if ad else ac,'zag_tax':None,'nett_type':nt,
                'component_count':len(grp),'component_descs':' | '.join(gross_rows['description'].tolist())})
            for i in idx_list: used.add(i)
    for idx,row in df.iterrows():
        if idx in used: continue
        r=row.to_dict()
        r.setdefault('netted',False); r.setdefault('zag_gross',None); r.setdefault('zag_tax',None)
        r.setdefault('nett_type',None); r.setdefault('component_count',1)
        r.setdefault('component_descs',r.get('description',''))
        netted_rows.append(r)
    return pd.DataFrame(netted_rows).reset_index(drop=True)

def _base(b,ccy,jref):
    return {'source':'ZAG','currency':ccy,'trade_date':b['trade_date'],
            'settlement_date':b['settlement_date'],'ref':b.get('ref'),
            'description':b['description'],'journal_ref':jref,
            'is_tax_line':False,'netted':True}

def _lastbal(grp):
    v=grp['balance'].dropna(); return v.iloc[-1] if len(v) else None

# ══════════════════════════════════════════════════════════════════════════════
# 3. MATCHING ENGINE
# ══════════════════════════════════════════════════════════════════════════════
_TYPE_MAP={'RECEIVE VS PAYMENT':['SETTLEMENT','TRADEID'],'DELIVER VS PAYMENT':['SETTLEMENT','TRADEID'],
           'CASH DEPOSIT':['TRANSFER FROM','WIRE IN','DEPOSIT'],'CASH WITHDRAWAL':['WIRE OUT','TRANSFER TO'],
           'CA DIVIDEND':['CASH DIVIDEND','DIVIDEND','DIV'],'CA INCOME':['CREDIT COUPON','COUPON','INTR']}
_INCOME_TYPES={'CA DIVIDEND','CA INCOME','CA INCOME (INTR)','CA PAY DUE'}
_INCOME_ZAG=('DIVIDEND','COUPON','INTR','INCOME','CREDIT COUPON')

def _clean(t):
    t=str(t or '').upper()
    for p in [r'\bARK CAPITAL MANAGEMENT \(DUBAI\)\b',r'\bARK CAPITAL MANAGEMENT\b',
              r'\bT24\b',r'\bSETTLEMENT OF TRADEID\b',r'[#\-_/\.]',r'\s+']:
        t=re.sub(p,' ',t)
    return t.strip()

def _score(fd,ft,zd,zr,zc=''):
    fk=_clean(fd+' '+ft); zk=_clean(str(zd)+' '+str(zr or '')+' '+str(zc or ''))
    ftk=_clean(ft)
    for key,hints in _TYPE_MAP.items():
        if key in ftk and any(h in zk for h in hints): return 'HIGH','type_map'
    fw={w for w in fk.split() if len(w)>3}; zw={w for w in zk.split() if len(w)>3}
    ov=fw&zw
    if len(ov)>=2: return 'HIGH',f'words:{",".join(sorted(ov)[:3])}'
    if len(ov)==1: return 'MEDIUM',f'words:{",".join(ov)}'
    return 'LOW','no_match'

def _tol(ft,zd):
    ft=str(ft or '').upper(); zd=str(zd or '').upper()
    return 15 if (any(t in ft for t in _INCOME_TYPES) or any(k in zd for k in _INCOME_ZAG)) else 5

def _signed(row):
    d=row.get('debit'); c=row.get('credit')
    if d is not None and not(isinstance(d,float) and np.isnan(d)): return -abs(float(d))
    if c is not None and not(isinstance(c,float) and np.isnan(c)): return abs(float(c))
    return 0.0

def reconcile(cust_df, zag_df):
    results=[]
    ccys=sorted(set(cust_df['currency'].dropna())|set(zag_df['currency'].dropna()))
    for ccy in ccys:
        cust=cust_df[cust_df['currency']==ccy].copy().reset_index(drop=True)
        zag=zag_df[zag_df['currency']==ccy].copy().reset_index(drop=True)
        cu=[False]*len(cust); zu=[False]*len(zag)
        # Pass 1: exact amount + date
        for ci,cr in cust.iterrows():
            ca=_signed(cr); cd=cr['settlement_date']
            for zi,zr in zag.iterrows():
                if zu[zi]: continue
                za=_signed(zr); zd=zr['settlement_date']
                if abs(abs(ca)-abs(za))<0.05 and pd.notna(cd) and pd.notna(zd) and pd.Timestamp(cd).date()==pd.Timestamp(zd).date():
                    conf,m=_score(cr.get('description',''),cr.get('txn_type',''),zr.get('description',''),zr.get('ref',''),zr.get('component_descs',''))
                    results.append(_mrow(ccy,cr,zr,ca,za,'MATCHED',conf,f'exact|{m}',''))
                    cu[ci]=True; zu[zi]=True; break
        # Pass 2: amount + tolerance
        for ci,cr in cust.iterrows():
            if cu[ci]: continue
            ca=_signed(cr); cd=cr.get('settlement_date') or cr.get('trade_date')
            for zi,zr in zag.iterrows():
                if zu[zi]: continue
                za=_signed(zr); zd=zr.get('settlement_date') or zr.get('trade_date')
                if abs(abs(ca)-abs(za))>=0.05: continue
                try: dd=abs((pd.Timestamp(cd)-pd.Timestamp(zd)).days) if pd.notna(cd) and pd.notna(zd) else 99
                except: dd=99
                tol=_tol(cr.get('txn_type',''),zr.get('description',''))
                if dd<=tol:
                    conf,m=_score(cr.get('description',''),cr.get('txn_type',''),zr.get('description',''),zr.get('ref',''),zr.get('component_descs',''))
                    lbl='income_delay' if tol==15 else 'settle_timing'
                    results.append(_mrow(ccy,cr,zr,ca,za,'MATCHED',conf,f'amt+{dd}d({lbl})|{m}',f'Date diff {dd}d'))
                    cu[ci]=True; zu[zi]=True; break
        for ci,cr in cust.iterrows():
            if not cu[ci]: results.append(_ucust(ccy,cr,_signed(cr)))
        for zi,zr in zag.iterrows():
            if not zu[zi]: results.append(_uzag(ccy,zr,_signed(zr)))
    return pd.DataFrame(results)

def _mrow(ccy,cr,zr,ca,za,st,conf,m,note):
    return dict(currency=ccy,match_status=st,confidence=conf,match_method=m,
                cust_settle_date=cr.get('settlement_date'),cust_type=cr.get('txn_type'),
                cust_description=cr.get('description'),cust_amount=round(ca,2),cust_balance=cr.get('balance'),
                custodian=cr.get('custodian',''),
                zag_date=zr.get('settlement_date'),zag_ref=zr.get('ref'),
                zag_description=zr.get('description'),zag_amount=round(za,2),zag_balance=zr.get('balance'),
                zag_journal=zr.get('journal_ref'),zag_netted=zr.get('netted',False),
                nett_type=zr.get('nett_type',''),zag_gross=zr.get('zag_gross'),zag_tax=zr.get('zag_tax'),
                difference=round(abs(ca)-abs(za),4),notes=note)

def _ucust(ccy,cr,ca):
    return dict(currency=ccy,match_status='UNMATCHED - CUSTODIAN ONLY',confidence='',match_method='',
                cust_settle_date=cr.get('settlement_date'),cust_type=cr.get('txn_type'),
                cust_description=cr.get('description'),cust_amount=round(ca,2),cust_balance=cr.get('balance'),
                custodian=cr.get('custodian',''),
                zag_date=None,zag_ref=None,zag_description=None,zag_amount=None,zag_balance=None,
                zag_journal=None,zag_netted=False,nett_type='',zag_gross=None,zag_tax=None,
                difference=None,notes='In custodian only — not in ZagTrader')

def _uzag(ccy,zr,za):
    return dict(currency=ccy,match_status='UNMATCHED - ZAG ONLY',confidence='',match_method='',
                cust_settle_date=None,cust_type=None,cust_description=None,cust_amount=None,
                cust_balance=None,custodian='',
                zag_date=zr.get('settlement_date'),zag_ref=zr.get('ref'),
                zag_description=zr.get('description'),zag_amount=round(za,2),zag_balance=zr.get('balance'),
                zag_journal=zr.get('journal_ref'),zag_netted=zr.get('netted',False),
                nett_type=zr.get('nett_type',''),zag_gross=zr.get('zag_gross'),zag_tax=zr.get('zag_tax'),
                difference=None,notes='In ZagTrader only — not in custodian')

def balance_summary(cust_bal,zag_bal):
    rows=[]
    for ccy in sorted(set(list(cust_bal)+list(zag_bal))):
        co=cust_bal.get(ccy,{}).get('starting'); cc=cust_bal.get(ccy,{}).get('ending')
        zo=zag_bal.get(ccy,{}).get('starting');  zc=zag_bal.get(ccy,{}).get('ending')
        diff=round(cc-zc,4) if cc is not None and zc is not None else None
        status='OK' if diff is not None and abs(diff)<0.05 else ('BREAK' if diff is not None else 'ONE SIDE MISSING')
        rows.append(dict(currency=ccy,cust_opening=co,cust_closing=cc,zag_opening=zo,zag_closing=zc,difference=diff,status=status))
    return pd.DataFrame(rows)

# ══════════════════════════════════════════════════════════════════════════════
# 4. REVIEW PARSER (read match IDs from annotated report)
# ══════════════════════════════════════════════════════════════════════════════
def parse_review(file_obj):
    if hasattr(file_obj,'read'): wb=load_workbook(io.BytesIO(file_obj.read()))
    else: wb=load_workbook(file_obj)
    if 'Unmatched Side-by-Side' not in wb.sheetnames: return []
    ws=wb['Unmatched Side-by-Side']
    # Detect header
    col_map={}
    for hr in [3,4,5]:
        for cell in ws[hr]:
            v=str(cell.value or '').strip().lower(); c=cell.column
            if v=='month' and c<=8 and 'fab_month' not in col_map: col_map['fab_month']=c
            elif v=='ccy' and c<=8 and 'fab_ccy' not in col_map: col_map['fab_ccy']=c
            elif any(x in v for x in ('cust date','fab date')): col_map['fab_date']=c
            elif any(x in v for x in ('cust type','fab type')): col_map['fab_type']=c
            elif any(x in v for x in ('cust description','fab description')): col_map['fab_desc']=c
            elif any(x in v for x in ('cust amount','fab amount')): col_map['fab_amt']=c
            elif 'match id' in v and c<=9: col_map['fab_match']=c
            elif v=='ccy' and c>=9 and 'zag_ccy' not in col_map: col_map['zag_ccy']=c
            elif 'zag date' in v: col_map['zag_date']=c
            elif 'zag ref' in v: col_map['zag_ref']=c
            elif 'zag description' in v: col_map['zag_desc']=c
            elif 'zag amount' in v: col_map['zag_amt']=c
            elif 'zag journal' in v: col_map['zag_jnl']=c
            elif 'match id' in v and c>=9: col_map['zag_match']=c
        if len(col_map)>=8: break
    def g(rd,k): return rd.get(col_map.get(k))
    def isn(v):
        try: float(v); return True
        except: return False
    data=[]
    for row in ws.iter_rows(min_row=6,max_row=ws.max_row):
        rd={cell.column:cell.value for cell in row}
        fm=g(rd,'fab_match'); zm=g(rd,'zag_match')
        if fm is not None and not isn(fm): fm=None
        if zm is not None and not isn(zm): zm=None
        if fm is None and zm is None: continue
        desc=g(rd,'fab_desc')
        if desc and 'Subtotal' in str(desc): continue
        data.append({'fab_ccy':str(g(rd,'fab_ccy')) if g(rd,'fab_ccy') else None,
                     'zag_ccy':str(g(rd,'zag_ccy')) if g(rd,'zag_ccy') else None,
                     'month_fab':g(rd,'fab_month'),'fab_date':g(rd,'fab_date'),
                     'fab_type':g(rd,'fab_type'),'fab_desc':g(rd,'fab_desc'),'fab_amt':g(rd,'fab_amt'),
                     'fab_match':int(float(fm)) if fm else None,
                     'zag_month':None,'zag_date':g(rd,'zag_date'),'zag_ref':g(rd,'zag_ref'),
                     'zag_desc':g(rd,'zag_desc'),'zag_amt':g(rd,'zag_amt'),'zag_jnl':g(rd,'zag_jnl'),
                     'zag_match':int(float(zm)) if zm else None})
    df=pd.DataFrame(data)
    if df.empty: return []
    fr=df[df['fab_match'].notna()&df['fab_ccy'].notna()].copy()
    zr=df[df['zag_match'].notna()&df['zag_ccy'].notna()].copy()
    fk=set(zip(fr['fab_ccy'],fr['fab_match'])); zk=set(zip(zr['zag_ccy'],zr['zag_match']))
    groups=[]
    for ccy,mid in sorted(fk|zk):
        fab_rows=fr[(fr['fab_ccy']==ccy)&(fr['fab_match']==mid)].to_dict('records')
        zag_rows=zr[(zr['zag_ccy']==ccy)&(zr['zag_match']==mid)].to_dict('records')
        groups.append({'ccy':ccy,'match_id':mid,'fab_rows':fab_rows,'zag_rows':zag_rows})
    return groups

def apply_manual(recon_df,groups):
    fab_sigs=set(); zag_sigs=set()
    manual_rows=[]
    for g in groups:
        for r in g.get('fab_rows',[]): 
            if r.get('fab_desc') and r.get('fab_amt') is not None:
                fab_sigs.add((str(r['fab_desc'])[:50],float(r['fab_amt'])))
        for r in g.get('zag_rows',[]):
            if r.get('zag_desc') and r.get('zag_amt') is not None:
                zag_sigs.add((str(r['zag_desc'])[:50],float(r['zag_amt'])))
        ft=sum(float(r['fab_amt']) for r in g.get('fab_rows',[]) if r.get('fab_amt') is not None)
        zt=sum(float(r['zag_amt']) for r in g.get('zag_rows',[]) if r.get('zag_amt') is not None)
        fb=g['fab_rows'][0] if g.get('fab_rows') else {}
        zb=g['zag_rows'][0] if g.get('zag_rows') else {}
        nf=len(g.get('fab_rows',[])); nz=len(g.get('zag_rows',[]))
        net=round(abs(ft)-abs(zt),2) if g.get('fab_rows') and g.get('zag_rows') else None
        flagged=net is not None and abs(net)>100 and abs(net)/max(abs(ft),abs(zt),0.01)>0.01
        notes=f"Match ID {g['match_id']} | {nf} cust | {nz} ZAG"
        if flagged: notes+=f" | ⚠ LARGE DIFF {abs(net):,.2f} — VERIFY"
        manual_rows.append({'currency':g['ccy'],'match_status':'MATCHED',
                            'confidence':'REVIEW REQUIRED' if flagged else 'HIGH',
                            'match_method':f"manual_id_{g['match_id']} ({nf}C:{nz}Z)",
                            'cust_settle_date':fb.get('fab_date'),'cust_type':fb.get('fab_type',''),
                            'cust_description':' + '.join(str(r.get('fab_desc',''))[:35] for r in g.get('fab_rows',[])),
                            'cust_amount':ft if g.get('fab_rows') else None,'cust_balance':None,'custodian':'',
                            'zag_date':zb.get('zag_date'),'zag_ref':zb.get('zag_ref',''),
                            'zag_description':' + '.join(str(r.get('zag_desc',''))[:35] for r in g.get('zag_rows',[])),
                            'zag_amount':zt if g.get('zag_rows') else None,'zag_balance':None,
                            'zag_journal':zb.get('zag_jnl',''),'zag_netted':False,'nett_type':'',
                            'zag_gross':None,'zag_tax':None,'difference':net,'notes':notes})
    def res(row):
        if row['match_status']=='UNMATCHED - CUSTODIAN ONLY':
            return (str(row.get('cust_description',''))[:50],float(row.get('cust_amount') or 0)) in fab_sigs
        return (str(row.get('zag_description',''))[:50],float(row.get('zag_amount') or 0)) in zag_sigs
    breaks=recon_df[recon_df['match_status'].str.startswith('UNMATCHED')].copy()
    breaks['_r']=breaks.apply(res,axis=1)
    remaining=breaks[~breaks['_r']].drop(columns=['_r'])
    matched=recon_df[recon_df['match_status']=='MATCHED']
    manual_df=pd.DataFrame(manual_rows)
    return pd.concat([matched,manual_df,remaining],ignore_index=True), manual_df

# ══════════════════════════════════════════════════════════════════════════════
# 5. REPORT BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_report(recon_df,balance_df,cust_df,zag_raw,zag_netted,groups,period,cust_name,manual_df=None):
    G=PatternFill("solid",fgColor="C6EFCE"); A=PatternFill("solid",fgColor="FFEB9C")
    R=PatternFill("solid",fgColor="FFC7CE"); Gr=PatternFill("solid",fgColor="F2F2F2")
    B=PatternFill("solid",fgColor="DEEAF1"); H=PatternFill("solid",fgColor="1F4E79")
    MH=PatternFill("solid",fgColor="2E4057"); MN=PatternFill("solid",fgColor="E2EFDA")
    LO=PatternFill("solid",fgColor="FCE4D6")
    RC=PatternFill("solid",fgColor="FCE4D6"); RZ=PatternFill("solid",fgColor="FFE4E1")
    HC=PatternFill("solid",fgColor="C00000"); HZ=PatternFill("solid",fgColor="833C00")
    HM=PatternFill("solid",fgColor="404040"); GA=PatternFill("solid",fgColor="F2F2F2")
    WH=PatternFill("solid",fgColor="FFFFFF")
    LG=PatternFill("solid",fgColor="EBF5E0"); LB=PatternFill("solid",fgColor="E0EDF5")
    hf=Font(bold=True,color="FFFFFF",size=10); n10=Font(size=10); n9=Font(size=9)
    def hr(ws,cols,row=1):
        for c,h in enumerate(cols,1):
            cell=ws.cell(row,c,h); cell.fill=H; cell.font=hf
            cell.alignment=Alignment(horizontal='center',wrap_text=True)
    def fd(v):
        if v is None: return ''
        try: ts=pd.Timestamp(v); return '' if pd.isnull(ts) else ts.strftime('%d/%m/%Y')
        except: return ''
    def fn(v):
        if v is None or (isinstance(v,float) and np.isnan(v)): return ''
        try: return round(float(v),2)
        except: return ''

    wb=Workbook()
    engine_matched=recon_df[recon_df['match_status']=='MATCHED'].copy()
    engine_breaks=recon_df[recon_df['match_status'].str.startswith('UNMATCHED')].copy()
    mH=len(engine_matched[engine_matched['confidence']=='HIGH'])
    mM=len(engine_matched[engine_matched['confidence']!='HIGH'])
    mMn=len(manual_df) if manual_df is not None else 0
    uC=len(engine_breaks[engine_breaks['match_status']=='UNMATCHED - CUSTODIAN ONLY'])
    uZ=len(engine_breaks[engine_breaks['match_status']=='UNMATCHED - ZAG ONLY'])
    tot=mH+mM+mMn+uC+uZ

    # Tab 1: Summary
    ws1=wb.active; ws1.title='Period Summary'
    ws1['A1']=f'Cash Reconciliation — {cust_name} vs ZagTrader  |  {period}'
    ws1['A1'].font=Font(bold=True,size=14,color='1F4E79')
    ws1['A2']=f'Generated: {datetime.today().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font=Font(size=10,color='595959'); ws1.append([])
    ws1.append(['MATCHING SUMMARY']); ws1.cell(ws1.max_row,1).font=Font(bold=True,size=11)
    ws1.append([])
    hr(ws1,['Metric','Count','% of Total'],row=ws1.max_row+1)
    for label,cnt,fill in [('Total items',tot,Gr),('Engine matched — high',mH,G),
                            ('Engine matched — medium',mM,A),('Manually confirmed',mMn,MN),
                            ('Unmatched — custodian only',uC,R if uC else G),
                            ('Unmatched — ZAG only',uZ,R if uZ else G)]:
        rn=ws1.max_row+1; pct=f'{round(cnt/tot*100,1) if tot else 0}%'
        for c,v in enumerate([label,cnt,pct],1): ws1.cell(rn,c,v).fill=fill; ws1.cell(rn,c).font=n10
    ws1.append([]); ws1.append([])
    ws1.append(['BALANCE COMPARISON']); ws1.cell(ws1.max_row,1).font=Font(bold=True,size=11)
    ws1.append([])
    hr(ws1,['Currency','Cust Opening','Cust Closing','ZAG Opening','ZAG Closing','Difference','Status'],row=ws1.max_row+1)
    for _,r in balance_df[~balance_df['status'].eq('ONE SIDE MISSING')].sort_values('currency').iterrows():
        rn=ws1.max_row+1; fill=G if r['status']=='OK' else R
        for c,v in enumerate([r['currency'],fn(r.get('cust_opening')),fn(r.get('cust_closing')),
                               fn(r.get('zag_opening')),fn(r.get('zag_closing')),fn(r.get('difference')),r['status']],1):
            cell=ws1.cell(rn,c,v); cell.fill=fill; cell.font=n10
            cell.alignment=Alignment(horizontal='right' if c in [2,3,4,5,6] else 'center')
    for c,w in enumerate([12,18,18,16,16,14,12],1): ws1.column_dimensions[get_column_letter(c)].width=w

    # Tab 2: Matched
    ws2=wb.create_sheet('Matched')
    hr(ws2,['CCY','Confidence','Method','Cust Date','Cust Type','Cust Description','Cust Amount',
            'ZAG Date','ZAG Ref','ZAG Description','ZAG Amount','ZAG Journal',
            'Netted?','Nett Type','ZAG Gross','ZAG WHT','Diff','Notes'])
    for _,r in engine_matched.iterrows():
        fill=G if r['confidence']=='HIGH' else A
        vals=[r['currency'],r['confidence'],r['match_method'],
              fd(r.get('cust_settle_date')),r.get('cust_type'),r.get('cust_description'),fn(r.get('cust_amount')),
              fd(r.get('zag_date')),r.get('zag_ref'),r.get('zag_description'),fn(r.get('zag_amount')),r.get('zag_journal'),
              'YES' if r.get('zag_netted') else '',r.get('nett_type',''),fn(r.get('zag_gross')),fn(r.get('zag_tax')),fn(r.get('difference')),r.get('notes','')]
        rn=ws2.max_row+1
        for c,v in enumerate(vals,1): cell=ws2.cell(rn,c,v); cell.fill=fill; cell.font=n9
    if manual_df is not None:
        for _,r in manual_df.iterrows():
            fill=LO if r.get('confidence')=='REVIEW REQUIRED' else MN
            vals=[r.get('currency'),r.get('confidence'),r.get('match_method'),
                  fd(r.get('cust_settle_date')),r.get('cust_type'),r.get('cust_description'),fn(r.get('cust_amount')),
                  fd(r.get('zag_date')),r.get('zag_ref'),r.get('zag_description'),fn(r.get('zag_amount')),r.get('zag_journal'),
                  '','','','',fn(r.get('difference')),r.get('notes','')]
            rn=ws2.max_row+1
            for c,v in enumerate(vals,1): cell=ws2.cell(rn,c,v); cell.fill=fill; cell.font=n9
    for c,w in enumerate([6,10,30,14,20,40,13,14,12,40,13,16,7,14,13,12,10,50],1): ws2.column_dimensions[get_column_letter(c)].width=w

    # Tab 3: Breaks
    ws3=wb.create_sheet('Breaks')
    ws3['A1']='UNMATCHED ITEMS'; ws3['A1'].font=Font(bold=True,size=12,color='C00000')
    ws3.append([])
    hr(ws3,['CCY','Status','Cust Date','Cust Type','Cust Description','Cust Amount',
            'ZAG Date','ZAG Ref','ZAG Description','ZAG Amount','ZAG Journal','Notes'],row=3)
    for _,r in engine_breaks.sort_values(['currency','match_status']).iterrows():
        note=str(r.get('notes',''))
        fill=A if any(n in note for n in ['Confirmed','Outside','duplicate','scope']) else R
        vals=[r['currency'],r['match_status'],fd(r.get('cust_settle_date')),r.get('cust_type'),
              r.get('cust_description'),fn(r.get('cust_amount')),fd(r.get('zag_date')),
              r.get('zag_ref'),r.get('zag_description'),fn(r.get('zag_amount')),r.get('zag_journal'),note]
        rn=ws3.max_row+1
        for c,v in enumerate(vals,1): cell=ws3.cell(rn,c,v); cell.fill=fill; cell.font=n9
    for c,w in enumerate([6,24,14,20,40,13,14,12,40,13,16,50],1): ws3.column_dimensions[get_column_letter(c)].width=w

    # Tab 4: Side-by-Side
    ws4=wb.create_sheet('Unmatched Side-by-Side')
    ws4['A1']='UNMATCHED SIDE-BY-SIDE — add Match IDs in col G (cust) and col P (ZAG), re-upload'
    ws4['A1'].font=Font(bold=True,size=11,color='1F4E79'); ws4.append([]); ws4.append([])
    HDR=3
    for c,h in enumerate(['Month','CCY','Cust Date','Cust Type','Cust Description','Cust Amount','Match ID','Notes'],1):
        cell=ws4.cell(HDR,c,h); cell.fill=HC; cell.font=hf; cell.alignment=Alignment(horizontal='center',wrap_text=True)
    ws4.cell(HDR,9,'').fill=HM
    for c,h in enumerate(['Month','CCY','ZAG Date','ZAG Ref','ZAG Description','ZAG Amount','ZAG Journal','Match ID','Notes'],10):
        cell=ws4.cell(HDR,c,h); cell.fill=HZ; cell.font=hf; cell.alignment=Alignment(horizontal='center',wrap_text=True)
    co=engine_breaks[engine_breaks['match_status']=='UNMATCHED - CUSTODIAN ONLY'].copy()
    zo=engine_breaks[engine_breaks['match_status']=='UNMATCHED - ZAG ONLY'].copy()
    co['_d']=pd.to_datetime(co['cust_settle_date'],errors='coerce')
    zo['_d']=pd.to_datetime(zo['zag_date'],errors='coerce')
    co=co.sort_values(['currency','_d']).reset_index(drop=True)
    zo=zo.sort_values(['currency','_d']).reset_index(drop=True)
    for ccy in sorted(set(co['currency'].unique())|set(zo['currency'].unique())):
        fc=co[co['currency']==ccy].reset_index(drop=True)
        zc=zo[zo['currency']==ccy].reset_index(drop=True)
        if len(fc)==0 and len(zc)==0: continue
        rn=ws4.max_row+1
        cell=ws4.cell(rn,1,f'Currency: {ccy}  —  Cust only: {len(fc)}  |  ZAG only: {len(zc)}')
        cell.fill=MH; cell.font=Font(bold=True,color='FFFFFF',size=10)
        ws4.merge_cells(start_row=rn,start_column=1,end_row=rn,end_column=18)
        for i in range(max(len(fc),len(zc))):
            rn=ws4.max_row+1; fa=GA if i%2==0 else WH
            if i<len(fc):
                r=fc.iloc[i]
                vf=['',r['currency'],fd(r.get('cust_settle_date')),str(r.get('cust_type') or '')[:25],str(r.get('cust_description') or '')[:55],fn(r.get('cust_amount')),'','']
                ff=RC
            else: vf=['','','','','','','','']; ff=fa
            for c,v in enumerate(vf,1): cell=ws4.cell(rn,c,v); cell.fill=ff; cell.font=n9
            ws4.cell(rn,9,'').fill=HM
            if i<len(zc):
                r=zc.iloc[i]
                vz=['',r['currency'],fd(r.get('zag_date')),str(r.get('zag_ref') or ''),str(r.get('zag_description') or '')[:55],fn(r.get('zag_amount')),str(r.get('zag_journal') or ''),'','']
                zf=RZ
            else: vz=['','','','','','','','','']; zf=fa
            for c,v in enumerate(vz,10): cell=ws4.cell(rn,c,v); cell.fill=zf; cell.font=n9
        rn=ws4.max_row+1
        ft=fc['cust_amount'].apply(lambda x: float(x) if x and not(isinstance(x,float) and np.isnan(x)) else 0).sum()
        zt=zc['zag_amount'].apply(lambda x: float(x) if x and not(isinstance(x,float) and np.isnan(x)) else 0).sum()
        ws4.cell(rn,5,f'Subtotal {ccy} Cust').font=Font(bold=True,size=9); ws4.cell(rn,5).alignment=Alignment(horizontal='right')
        ws4.cell(rn,6,round(ft,2)).font=Font(bold=True,size=9); ws4.cell(rn,9,'').fill=HM
        ws4.cell(rn,14,f'Subtotal {ccy} ZAG').font=Font(bold=True,size=9); ws4.cell(rn,14).alignment=Alignment(horizontal='right')
        ws4.cell(rn,15,round(zt,2)).font=Font(bold=True,size=9); ws4.append([])
    for c,w in enumerate([10,6,12,20,45,14,10,22,2,10,6,12,14,45,14,16,10,22],1): ws4.column_dimensions[get_column_letter(c)].width=w
    ws4.freeze_panes='A4'

    # Tab 5: ZAG Netting
    ws5=wb.create_sheet('ZAG Netting Detail')
    ws5['A1']='ZagTrader: Multi-line journal entries collapsed before matching'; ws5['A1'].font=Font(bold=True,size=11,color='1F4E79')
    hr(ws5,['Currency','Date','Journal Ref','Nett Type','Description','Gross Amt','WHT Tax','Net Amt','Component Lines'],row=3)
    for _,r in zag_netted[zag_netted['netted']==True].iterrows():
        d=r.get('debit') or 0; cr=r.get('credit') or 0; rn=ws5.max_row+1
        for c,v in enumerate([r['currency'],fd(r['trade_date']),r['journal_ref'],r.get('nett_type','gross+tax'),
                               r['description'],fn(r.get('zag_gross')),fn(r.get('zag_tax')),fn(d if d else cr),
                               str(r.get('component_descs',''))[:120]],1):
            cell=ws5.cell(rn,c,v); cell.fill=B; cell.font=n9
    for c,w in enumerate([8,12,16,18,40,14,12,16,120],1): ws5.column_dimensions[get_column_letter(c)].width=w

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# 6. STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("⚙️ Configuration")
    st.markdown("---")
    st.subheader("Period")
    period_label=st.text_input("Period label",value=datetime.today().strftime('%b %Y'))
    st.markdown("---")
    st.subheader("Matching tolerances")
    income_tol=st.slider("Income/dividend delay (days)",5,30,15)
    settle_tol=st.slider("Settlement tolerance (days)",1,10,5)
    st.markdown("---")
    st.caption("Custodian: FAB (any .xls file)")
    st.caption("To add custodians: contact your developer")

st.title("🔄 Cash Reconciliation Agent")
st.caption("Upload custodian statement(s) + ZagTrader export → download reconciliation report")

tab_run,tab_review,tab_help=st.tabs(["▶ Run reconciliation","📋 Apply manual matches","❓ How to use"])

with tab_run:
    col_left,col_right=st.columns(2)
    with col_left:
        st.subheader("📁 Custodian statement(s)")
        cust_files=st.file_uploader("Upload FAB .xls file(s)",type=["xls","xlsx"],accept_multiple_files=True,key="cust")
        if cust_files:
            for f in cust_files:
                if f.name.lower().endswith('.xls'): st.success(f"✓ {f.name}  →  FAB parser")
                else: st.warning(f"⚠ {f.name}  →  unknown format")
    with col_right:
        st.subheader("📁 ZagTrader export")
        zag_file=st.file_uploader("Upload ZagTrader .xlsx file",type=["xlsx"],key="zag")
        if zag_file: st.success(f"✓ {zag_file.name}")
    st.markdown("---")
    st.subheader("📋 Previous review file (optional)")
    st.caption("Upload a previously downloaded report where you added Match IDs in the Side-by-Side sheet.")
    prev_files=st.file_uploader("Previous report(s) with match IDs",type=["xlsx"],accept_multiple_files=True,key="prev")
    st.markdown("---")
    run_btn=st.button("▶ Run reconciliation",type="primary",disabled=(not cust_files or not zag_file))

    if run_btn and cust_files and zag_file:
        with st.spinner("Running reconciliation..."):
            try:
                # Parse custodian files
                all_dfs=[]; all_bals={}
                for f in cust_files:
                    if not f.name.lower().endswith('.xls'):
                        st.warning(f"Skipped {f.name}"); continue
                    tmp=f"/tmp/{f.name}"
                    with open(tmp,"wb") as t: t.write(f.read())
                    df,bal=parse_fab(tmp)
                    df=df[df['currency']!='ZAR'].reset_index(drop=True)
                    all_dfs.append(df)
                    for ccy,info in bal.items():
                        ex=all_bals.get(ccy,{})
                        if 'starting' not in ex: ex['starting']=info.get('starting')
                        ex['ending']=info.get('ending'); all_bals[ccy]=ex
                if not all_dfs: st.error("No FAB files could be parsed."); st.stop()
                cust_df=pd.concat(all_dfs,ignore_index=True)

                # Parse ZAG
                tmp_zag="/tmp/zag_upload.xlsx"
                with open(tmp_zag,"wb") as t: t.write(zag_file.read())
                zag_raw,zag_bal=parse_zag_raw(tmp_zag)
                zag_netted=net_zag(zag_raw)

                # Reconcile
                recon_df=reconcile(cust_df,zag_netted)
                balance_df=balance_summary(all_bals,zag_bal)

                # Apply manual matches
                all_groups=[]; manual_df=None
                for pf in (prev_files or []):
                    pf.seek(0); groups=parse_review(pf)
                    offset=len(all_groups)*1000
                    for g in groups: g['match_id']=g['match_id']+offset
                    all_groups.extend(groups)
                if all_groups:
                    recon_df,manual_df=apply_manual(recon_df,all_groups)

                # Stats
                matched=recon_df[recon_df['match_status']=='MATCHED']
                breaks=recon_df[recon_df['match_status'].str.startswith('UNMATCHED')]
                total=len(recon_df)
                rate=round(len(matched)/total*100,1) if total else 0

                m1,m2,m3,m4,m5=st.columns(5)
                m1.metric("Total items",total)
                m2.metric("Matched",len(matched),delta=f"{rate}%")
                m3.metric("Breaks",len(breaks),delta_color="inverse")
                m4.metric("Bal. OK",len(balance_df[balance_df['status']=='OK']))
                m5.metric("Bal. breaks",len(balance_df[balance_df['status']=='BREAK']),delta_color="inverse")
                st.markdown("---")

                # Balance preview
                st.subheader("Balance comparison")
                bd=balance_df[~balance_df['status'].eq('ONE SIDE MISSING')].copy()
                def cs(v): return 'background-color:#C6EFCE' if v=='OK' else ('background-color:#FFC7CE' if v=='BREAK' else '')
                st.dataframe(bd[['currency','cust_opening','cust_closing','zag_opening','zag_closing','difference','status']]
                             .rename(columns={'currency':'CCY','cust_opening':'Cust Open','cust_closing':'Cust Close',
                                              'zag_opening':'ZAG Open','zag_closing':'ZAG Close','difference':'Diff','status':'Status'})
                             .style.map(cs,subset=["Status"]),use_container_width=True,hide_index=True)

                # Breaks preview
                if len(breaks)>0:
                    st.subheader(f"Top unmatched items ({min(15,len(breaks))} of {len(breaks)})")
                    pc=[c for c in ['currency','match_status','cust_settle_date','cust_description','cust_amount','zag_date','zag_description','zag_amount','notes'] if c in breaks.columns]
                    st.dataframe(breaks[pc].head(15),use_container_width=True,hide_index=True)

                # Build and download report
                report=build_report(recon_df,balance_df,cust_df,zag_raw,zag_netted,all_groups,period_label,'FAB',manual_df)
                fname=f"Recon_{period_label.replace(' ','_')}_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button("⬇ Download reconciliation report (.xlsx)",data=report,file_name=fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
                st.success(f"✓ Done — {len(matched)} matched, {len(breaks)} breaks, {rate}% match rate")
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback; st.code(traceback.format_exc())

with tab_review:
    st.subheader("Apply manual match IDs from a reviewed report")
    st.markdown("""
    1. Download a report from the **Run** tab
    2. Open the **Unmatched Side-by-Side** sheet
    3. Add the same number in column **G** (custodian) and column **P** (ZAG) to link matching transactions
    4. Save and upload the file below, then re-run
    """)
    rv=st.file_uploader("Upload annotated report",type=["xlsx"],key="rv")
    if rv:
        rv.seek(0); groups=parse_review(rv)
        st.success(f"✓ Found {len(groups)} match groups")
        if groups:
            rows=[]
            for g in groups:
                ft=sum(float(r['fab_amt']) for r in g.get('fab_rows',[]) if r.get('fab_amt') is not None)
                zt=sum(float(r['zag_amt']) for r in g.get('zag_rows',[]) if r.get('zag_amt') is not None)
                net=round(ft+zt,2)
                flagged=abs(net)>100 and abs(net)/max(abs(ft),abs(zt),0.01)>0.01
                rows.append({'CCY':g['ccy'],'ID':g['match_id'],'Cust lines':len(g.get('fab_rows',[])),'ZAG lines':len(g.get('zag_rows',[])),'Cust total':round(ft,2),'ZAG total':round(zt,2),'Net diff':net,'Flag':'⚠ VERIFY' if flagged else '✓'})
            st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

with tab_help:
    st.subheader("How to use")
    st.markdown("""
    **Run a reconciliation:**
    1. Upload your FAB `.xls` file(s) on the left
    2. Upload your ZagTrader `.xlsx` file on the right
    3. Set the period label (e.g. `Jan 2026`)
    4. Click **▶ Run reconciliation**
    5. Download the Excel report

    **Manual matching (for items the engine couldn't match):**
    1. Open the downloaded report → go to the **Unmatched Side-by-Side** tab
    2. Find a custodian row (left, salmon) and its matching ZAG row (right, pink)
    3. Type the same number in column **G** (custodian side) and column **P** (ZAG side)
    4. For one FAB → multiple ZAG: give all ZAG rows the same ID as the FAB row
    5. Save the file → upload it in the **Apply manual matches** tab → re-run

    **Matching logic:**
    - ZAG gross income + WHT tax → netted to match custodian net posting
    - Coupon notional splits → aggregated
    - Transfer + bank charge → collapsed to principal
    - Income/dividend date tolerance: **15 days**
    - Settlement tolerance: **5 days**
    """)
